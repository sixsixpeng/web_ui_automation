# -*- coding: utf-8 -*-
"""
data_util - 多格式数据处理工具
功能：统一封装 YAML/Excel/CSV 读写、参数化数据源加载
"""

from typing import Any, List

from utils.crypto_util import crypto_util
from utils.log_util import get_logger
from utils.path_util import path_util
from utils.yaml_json_util import yaml_json_util

logger = get_logger("data_util")


class DataUtil:
    """多格式数据读取工具（单例）"""

    _instance = None

    def __new__(cls):
        """Singleton pattern for DataUtil"""
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        """Init data util"""
        if self._initialized:
            return
        self._initialized = True
        self._cache = {}
        self._env_config = None
        self._user_accounts = None
        self._case_data = None
        logger.info("DataUtil 初始化完成")

    # ========== YAML 数据读取 ==========

    def get_env_config(self) -> dict:
        """获取环境配置（env_config.yaml）"""
        if self._env_config is None:
            file_path = str(path_util.data_dir / "env_config.yaml")
            self._env_config = yaml_json_util.load_yaml(file_path)
            logger.info("环境配置已加载")
        return self._env_config

    def get_user_accounts(self, decrypt: bool = True) -> dict:
        """获取用户账号配置（自动解密）"""
        if self._user_accounts is None:
            file_path = str(path_util.data_dir / "user_account.yaml")
            raw_data = yaml_json_util.load_yaml(file_path)
            if decrypt:
                self._user_accounts = crypto_util._decrypt_dict(raw_data)
            else:
                self._user_accounts = raw_data
            logger.info("用户账号已加载")
        return self._user_accounts

    def get_case_data(self) -> dict:
        """获取用例业务数据"""
        if self._case_data is None:
            file_path = str(path_util.data_dir / "case_data.yaml")
            self._case_data = yaml_json_util.load_yaml(file_path)
            logger.info("用例数据已加载")
        return self._case_data

    def get_yaml(self, filename: str) -> dict:
        """加载指定 YAML 文件"""
        file_path = str(path_util.data_dir / filename)
        return yaml_json_util.load_yaml(file_path)

    # ========== Excel 数据读取 ==========

    def read_excel(self, filename: str, sheet_name: str = None) -> List[dict]:
        """读取 Excel 文件，返回字典列表"""
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl 未安装，请执行: pip install openpyxl")
            raise

        file_path = str(path_util.data_dir / filename)
        if not path_util.exists(file_path):
            logger.error(f"Excel not found: {file_path}")
            raise FileNotFoundError(f"Excel not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h) if h is not None else "" for h in rows[0]]
        data = []
        for row in rows[1:]:
            if any(cell is not None for cell in row):
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = cell
                data.append(row_dict)

        wb.close()
        logger.debug(f"Excel 读取完成: {filename}, {len(data)} 条数据")
        return data

    def write_excel(self, filename: str, data: List[dict], sheet_name: str = "Sheet1"):
        """写入 Excel 文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl 未安装")
            raise

        file_path = str(path_util.data_dir / filename)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name

        if data:
            # 写入表头
            headers = list(data[0].keys())
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        wb.save(file_path)
        logger.info(f"Excel 写入完成: {filename}")

    # ========== CSV 数据读取 ==========

    def read_csv(self, filename: str) -> List[dict]:
        """读取 CSV 文件"""
        import csv
        file_path = str(path_util.data_dir / filename)
        if not path_util.exists(file_path):
            logger.error(f"CSV not found: {file_path}")
            raise FileNotFoundError(f"CSV not found: {file_path}")

        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            data = list(reader)

        logger.debug(f"CSV 读取完成: {filename}, {len(data)} 条数据")
        return data

    def write_csv(self, filename: str, data: List[dict]):
        """写入 CSV 文件"""
        import csv
        file_path = str(path_util.data_dir / filename)
        if not data:
            return

        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)

        logger.info(f"CSV 写入完成: {filename}")

    # ========== 便捷获取方法 ==========

    def get_login_valid_user(self) -> dict:
        """获取有效的登录用户数据"""
        accounts = self.get_user_accounts()
        return accounts.get("valid_user", {})

    def get_login_invalid_user(self) -> dict:
        """获取无效的登录用户数据"""
        accounts = self.get_user_accounts()
        return accounts.get("invalid_user", {})

    def get_case_expect(self, key: str) -> Any:
        """获取预期结果数据"""
        case_data = self.get_case_data()
        return yaml_json_util.get_nested_value(case_data, key)

    def parametrize_from_excel(self, filename: str, sheet_name: str = None) -> List[tuple]:
        """从 Excel 读取参数化数据"""
        data = self.read_excel(filename, sheet_name)
        if not data:
            return []
        return [tuple(d.values()) for d in data]

    def parametrize_from_csv(self, filename: str) -> List[tuple]:
        """从 CSV 读取参数化数据"""
        data = self.read_csv(filename)
        if not data:
            return []
        return [tuple(d.values()) for d in data]


# 全局单例
data_util = DataUtil()
